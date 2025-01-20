import cv2
from pyzbar.pyzbar import decode
import pandas as pd
from datetime import datetime
import numpy as np
from openpyxl.styles import PatternFill

class MultiQRCodeScanner:
    def __init__(self, camera_index=0, excel_path="multi_qr_codes.xlsx"):
        self.camera_index = camera_index
        self.excel_path = excel_path
        self.captured_codes = {}  # Changed to dict to track count
        self.qr_data = []
        self.cap = None
        self.frame_count = 0
        self.last_save_time = datetime.now()
        
    def initialize_camera(self):
        """Initialize the camera capture"""
        self.cap = cv2.VideoCapture(self.camera_index)
        if not self.cap.isOpened():
            raise Exception("Could not open camera")
            
        # Set camera resolution if needed
        self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, 1920)
        self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 1080)
        
    def draw_qr_info(self, frame, decoded_objects):
        """Draw QR code information on frame"""
        for idx, obj in enumerate(decoded_objects):
            # Get the QR code points
            points = obj.polygon
            if points:
                # Draw the QR code boundary
                pts = np.array([(p.x, p.y) for p in points], np.int32)
                pts = pts.reshape((-1, 1, 2))
                
                # Check if this is a duplicate QR code
                try:
                    data = obj.data.decode('utf-8')
                    color = (0, 0, 255) if data in self.captured_codes else (0, 255, 0)  # Red if duplicate, green if new
                    cv2.polylines(frame, [pts], True, color, 2)
                    
                    # Add a label with QR code content
                    x = points[0].x
                    y = points[0].y
                    truncated_data = data[:20] + "..." if len(data) > 20 else data
                    cv2.putText(frame, f"QR {idx+1}: {truncated_data}", 
                              (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 
                              0.5, color, 2)
                except:
                    pass
                
        # Add counter for total QR codes in frame
        cv2.putText(frame, f"QR Codes in frame: {len(decoded_objects)}", 
                   (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 
                   1, (0, 255, 0), 2)
        
        # Add counter for unique codes captured
        cv2.putText(frame, f"Total unique codes: {len(self.captured_codes)}", 
                   (10, 60), cv2.FONT_HERSHEY_SIMPLEX, 
                   1, (0, 255, 0), 2)
        
        return frame
        
    def process_frame(self, frame):
        """Process a single frame and detect QR codes"""
        self.frame_count += 1
        decoded_objects = decode(frame)
        codes_in_frame = 0
        
        for obj in decoded_objects:
            try:
                data = obj.data.decode('utf-8')
                codes_in_frame += 1
                
                # Track if this is a duplicate
                is_duplicate = data in self.captured_codes
                if data in self.captured_codes:
                    self.captured_codes[data] += 1
                else:
                    self.captured_codes[data] = 1
                
                self.qr_data.append({
                    'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'frame_number': self.frame_count,
                    'qr_content': data,
                    'qr_type': obj.type,
                    'status': 'duplicate' if is_duplicate else 'new',
                    'scan_count': self.captured_codes[data]
                })
                
                print(f"{'Duplicate' if is_duplicate else 'New'} QR Code detected: {data}")
                    
            except Exception as e:
                print(f"Error processing QR code: {str(e)}")
        
        # Draw information on frame
        frame = self.draw_qr_info(frame, decoded_objects)
        
        # Save to Excel if codes were detected in this frame
        if codes_in_frame > 0:
            self.save_to_excel()
            
        return frame
    
    def save_to_excel(self):
        """Save the captured QR code data to Excel with formatting"""
        try:
            # Create DataFrame
            df = pd.DataFrame(self.qr_data)
            
            # Save to Excel
            with pd.ExcelWriter(self.excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
                
                # Access the workbook and sheet
                workbook = writer.book
                worksheet = writer.sheets['Sheet1']
                
                # Define fill patterns
                red_fill = PatternFill(start_color='FFFF0000',
                                     end_color='FFFF0000',
                                     fill_type='solid')
                
                # Apply conditional formatting
                for idx, row in enumerate(df['status'], start=2):  # start=2 because excel uses 1-based index and we have header
                    if row == 'duplicate':
                        for col in range(1, len(df.columns) + 1):
                            worksheet.cell(row=idx, column=col).fill = red_fill
                
            print(f"Excel file updated with {len(self.qr_data)} total records")
        except Exception as e:
            print(f"Error saving to Excel: {str(e)}")
    
    def start_scanning(self):
        """Start the scanning process"""
        try:
            self.initialize_camera()
            print("Starting multi QR code scanning...")
            print("Press 'q' to quit")
            
            while True:
                ret, frame = self.cap.read()
                if not ret:
                    print("Failed to grab frame")
                    break
                
                # Process the frame
                processed_frame = self.process_frame(frame)
                
                # Display the frame
                cv2.imshow('Multi QR Code Scanner', processed_frame)
                
                # Break the loop if 'q' is pressed
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    break
                
        except Exception as e:
            print(f"Error during scanning: {str(e)}")
        
        finally:
            if self.cap is not None:
                self.cap.release()
            cv2.destroyAllWindows()
            self.save_to_excel()  # Final save

def main():
    scanner = MultiQRCodeScanner(
        camera_index=0,  # Change this if you have multiple cameras
        excel_path="multi_qr_codes.xlsx"
    )
    scanner.start_scanning()

if __name__ == "__main__":
    main()